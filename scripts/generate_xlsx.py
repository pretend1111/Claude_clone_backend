#!/usr/bin/env python3.11
"""Generate XLSX files from JSON input via stdin.

Input JSON format:
{
  "outputPath": "/path/to/output.xlsx",
  "title": "Document Title",
  "sheets": [
    {
      "name": "Sheet1",
      "headers": ["Col A", "Col B"],
      "rows": [["val1", "val2"], ...],
      "columnWidths": [20, 30],
      "formulas": [{"cell": "C2", "formula": "=SUM(A2:B2)"}]
    }
  ]
}
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def auto_width(ws, headers, rows, explicit_widths):
    """Set column widths: use explicit if provided, otherwise auto-fit."""
    for col_idx in range(1, len(headers) + 1):
        if explicit_widths and col_idx - 1 < len(explicit_widths) and explicit_widths[col_idx - 1]:
            ws.column_dimensions[get_column_letter(col_idx)].width = explicit_widths[col_idx - 1]
        else:
            max_len = len(str(headers[col_idx - 1])) if col_idx - 1 < len(headers) else 8
            for row in rows:
                if col_idx - 1 < len(row):
                    cell_len = len(str(row[col_idx - 1])) if row[col_idx - 1] is not None else 0
                    max_len = max(max_len, cell_len)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)


def build_sheet(ws, sheet_data):
    """Populate a worksheet from sheet data dict."""
    headers = sheet_data.get("headers", [])
    rows = sheet_data.get("rows", [])
    col_widths = sheet_data.get("columnWidths")
    formulas = sheet_data.get("formulas", [])

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    # Write headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=coerce_value(val))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Apply formulas
    for f in formulas:
        cell_ref = f.get("cell", "")
        formula_str = f.get("formula", "")
        if cell_ref and formula_str:
            ws[cell_ref] = formula_str

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-width
    auto_width(ws, headers, rows, col_widths)


def coerce_value(val):
    """Try to convert string values to numbers where appropriate."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return s
        # Percentage
        if s.endswith("%"):
            try:
                return float(s[:-1]) / 100
            except ValueError:
                pass
        # Number
        try:
            if "." in s:
                return float(s)
            return int(s)
        except ValueError:
            pass
    return val


def main():
    data = json.load(sys.stdin)
    output_path = data["outputPath"]
    sheets = data.get("sheets", [])

    if not sheets:
        print(json.dumps({"error": "No sheets provided"}))
        sys.exit(1)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_data in sheets:
        name = sheet_data.get("name", "Sheet")[:31]  # Excel 31 char limit
        ws = wb.create_sheet(title=name)
        build_sheet(ws, sheet_data)

    wb.save(output_path)
    print(json.dumps({"success": True, "path": output_path}))


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)
